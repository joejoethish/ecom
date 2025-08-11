"""
Chart services for data processing and business logic
"""
import json
import hashlib
import io
import base64
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional
from django.core.cache import cache
from django.db.models import Q, Count, Sum, Avg, Min, Max
from django.utils import timezone
from django.conf import settings
import pandas as pd
import numpy as np
from PIL import Image, ImageDraw
import matplotlib.pyplot as plt
import seaborn as sns
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
import openpyxl
from openpyxl.chart import LineChart, BarChart, PieChart

from .chart_models import Chart, ChartDataCache, ChartPerformanceMetric
from apps.orders.models import Order
from apps.products.models import Product
from apps.customers.models import Customer

class ChartDataService:
    """Service for fetching and processing chart data"""
    
    def __init__(self):
        self.cache_timeout = getattr(settings, 'CHART_CACHE_TIMEOUT', 300)
    
    def get_chart_data(self, chart: Chart, filters: Dict = None, use_cache: bool = True) -> Dict:
        """Get data for a specific chart"""
        cache_key = self._generate_cache_key(chart, filters)
        
        if use_cache:
            cached_data = cache.get(cache_key)
            if cached_data:
                return cached_data
        
        # Fetch data based on chart type and data source
        data = self._fetch_data(chart, filters)
        
        # Process data for chart format
        processed_data = self._process_data(chart, data)
        
        # Cache the result
        if use_cache:
            cache.set(cache_key, processed_data, self.cache_timeout)
        
        return processed_data
    
    def _generate_cache_key(self, chart: Chart, filters: Dict = None) -> str:
        """Generate cache key for chart data"""
        key_data = {
            'chart_id': str(chart.id),
            'data_source': chart.data_source,
            'filters': filters or {},
            'config': chart.config
        }
        key_string = json.dumps(key_data, sort_keys=True)
        return f"chart_data_{hashlib.md5(key_string.encode()).hexdigest()}"
    
    def _fetch_data(self, chart: Chart, filters: Dict = None) -> List[Dict]:
        """Fetch raw data based on chart data source"""
        data_source = chart.data_source
        filters = filters or {}
        
        if data_source == 'sales_overview':
            return self._get_sales_data(filters)
        elif data_source == 'product_performance':
            return self._get_product_performance_data(filters)
        elif data_source == 'customer_analytics':
            return self._get_customer_analytics_data(filters)
        elif data_source == 'inventory_levels':
            return self._get_inventory_data(filters)
        elif data_source == 'revenue_trends':
            return self._get_revenue_trends_data(filters)
        else:
            return []
    
    def _get_sales_data(self, filters: Dict) -> List[Dict]:
        """Get sales data"""
        queryset = Order.objects.all()
        
        # Apply date filters
        if filters.get('date_from'):
            queryset = queryset.filter(created_at__gte=filters['date_from'])
        if filters.get('date_to'):
            queryset = queryset.filter(created_at__lte=filters['date_to'])
        
        # Group by date
        if filters.get('group_by') == 'day':
            data = queryset.extra(
                select={'date': "DATE(created_at)"}
            ).values('date').annotate(
                total_orders=Count('id'),
                total_revenue=Sum('total_amount')
            ).order_by('date')
        elif filters.get('group_by') == 'month':
            data = queryset.extra(
                select={'date': "DATE_FORMAT(created_at, '%%Y-%%m')"}
            ).values('date').annotate(
                total_orders=Count('id'),
                total_revenue=Sum('total_amount')
            ).order_by('date')
        else:
            data = queryset.aggregate(
                total_orders=Count('id'),
                total_revenue=Sum('total_amount')
            )
            return [data]
        
        return list(data)
    
    def _get_product_performance_data(self, filters: Dict) -> List[Dict]:
        """Get product performance data"""
        from apps.orders.models import OrderItem
        
        queryset = OrderItem.objects.select_related('product')
        
        if filters.get('date_from'):
            queryset = queryset.filter(order__created_at__gte=filters['date_from'])
        if filters.get('date_to'):
            queryset = queryset.filter(order__created_at__lte=filters['date_to'])
        
        data = queryset.values(
            'product__name'
        ).annotate(
            total_sold=Sum('quantity'),
            total_revenue=Sum('price')
        ).order_by('-total_revenue')[:10]
        
        return list(data)
    
    def _get_customer_analytics_data(self, filters: Dict) -> List[Dict]:
        """Get customer analytics data"""
        queryset = Customer.objects.all()
        
        # Customer registration trends
        if filters.get('metric') == 'registrations':
            if filters.get('group_by') == 'month':
                data = queryset.extra(
                    select={'date': "DATE_FORMAT(created_at, '%%Y-%%m')"}
                ).values('date').annotate(
                    count=Count('id')
                ).order_by('date')
            else:
                data = queryset.extra(
                    select={'date': "DATE(created_at)"}
                ).values('date').annotate(
                    count=Count('id')
                ).order_by('date')
        else:
            # Customer lifetime value
            data = queryset.annotate(
                total_orders=Count('order'),
                total_spent=Sum('order__total_amount')
            ).values('total_spent').annotate(
                customer_count=Count('id')
            )
        
        return list(data)
    
    def _get_inventory_data(self, filters: Dict) -> List[Dict]:
        """Get inventory data"""
        queryset = Product.objects.all()
        
        if filters.get('low_stock_only'):
            queryset = queryset.filter(stock_quantity__lt=10)
        
        data = queryset.values(
            'name', 'stock_quantity', 'category__name'
        ).order_by('stock_quantity')
        
        return list(data)
    
    def _get_revenue_trends_data(self, filters: Dict) -> List[Dict]:
        """Get revenue trends data"""
        queryset = Order.objects.filter(status='completed')
        
        # Apply date filters
        if filters.get('date_from'):
            queryset = queryset.filter(created_at__gte=filters['date_from'])
        if filters.get('date_to'):
            queryset = queryset.filter(created_at__lte=filters['date_to'])
        
        # Group by period
        period = filters.get('period', 'day')
        if period == 'hour':
            data = queryset.extra(
                select={'period': "DATE_FORMAT(created_at, '%%Y-%%m-%%d %%H:00')"}
            ).values('period').annotate(
                revenue=Sum('total_amount')
            ).order_by('period')
        elif period == 'day':
            data = queryset.extra(
                select={'period': "DATE(created_at)"}
            ).values('period').annotate(
                revenue=Sum('total_amount')
            ).order_by('period')
        elif period == 'week':
            data = queryset.extra(
                select={'period': "YEARWEEK(created_at)"}
            ).values('period').annotate(
                revenue=Sum('total_amount')
            ).order_by('period')
        else:  # month
            data = queryset.extra(
                select={'period': "DATE_FORMAT(created_at, '%%Y-%%m')"}
            ).values('period').annotate(
                revenue=Sum('total_amount')
            ).order_by('period')
        
        return list(data)
    
    def _process_data(self, chart: Chart, raw_data: List[Dict]) -> Dict:
        """Process raw data into chart format"""
        chart_type = chart.chart_type
        config = chart.config
        
        if chart_type == 'line':
            return self._process_line_chart_data(raw_data, config)
        elif chart_type == 'bar':
            return self._process_bar_chart_data(raw_data, config)
        elif chart_type == 'pie':
            return self._process_pie_chart_data(raw_data, config)
        elif chart_type == 'area':
            return self._process_area_chart_data(raw_data, config)
        else:
            return self._process_generic_chart_data(raw_data, config)
    
    def _process_line_chart_data(self, data: List[Dict], config: Dict) -> Dict:
        """Process data for line charts"""
        labels = []
        datasets = []
        
        if not data:
            return {'labels': [], 'datasets': []}
        
        # Extract labels (x-axis)
        label_field = config.get('label_field', 'date')
        value_fields = config.get('value_fields', ['total_revenue'])
        
        for item in data:
            if label_field in item:
                labels.append(str(item[label_field]))
        
        # Create datasets for each value field
        for field in value_fields:
            dataset_data = []
            for item in data:
                dataset_data.append(float(item.get(field, 0) or 0))
            
            datasets.append({
                'label': field.replace('_', ' ').title(),
                'data': dataset_data,
                'borderColor': config.get('colors', ['#3B82F6'])[len(datasets) % len(config.get('colors', ['#3B82F6']))],
                'backgroundColor': config.get('colors', ['#3B82F6'])[len(datasets) % len(config.get('colors', ['#3B82F6']))] + '20',
                'tension': 0.4
            })
        
        return {
            'labels': labels,
            'datasets': datasets,
            'metadata': {
                'total_points': len(labels),
                'chart_type': 'line'
            }
        }
    
    def _process_bar_chart_data(self, data: List[Dict], config: Dict) -> Dict:
        """Process data for bar charts"""
        labels = []
        datasets = []
        
        if not data:
            return {'labels': [], 'datasets': []}
        
        label_field = config.get('label_field', 'name')
        value_field = config.get('value_field', 'total_revenue')
        
        dataset_data = []
        for item in data:
            if label_field in item:
                labels.append(str(item[label_field]))
                dataset_data.append(float(item.get(value_field, 0) or 0))
        
        datasets.append({
            'label': value_field.replace('_', ' ').title(),
            'data': dataset_data,
            'backgroundColor': config.get('colors', ['#3B82F6']),
            'borderColor': config.get('border_colors', ['#1E40AF']),
            'borderWidth': 1
        })
        
        return {
            'labels': labels,
            'datasets': datasets,
            'metadata': {
                'total_items': len(labels),
                'chart_type': 'bar'
            }
        }
    
    def _process_pie_chart_data(self, data: List[Dict], config: Dict) -> Dict:
        """Process data for pie charts"""
        labels = []
        dataset_data = []
        
        if not data:
            return {'labels': [], 'datasets': []}
        
        label_field = config.get('label_field', 'name')
        value_field = config.get('value_field', 'count')
        
        for item in data:
            if label_field in item:
                labels.append(str(item[label_field]))
                dataset_data.append(float(item.get(value_field, 0) or 0))
        
        datasets = [{
            'data': dataset_data,
            'backgroundColor': config.get('colors', [
                '#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0', 
                '#9966FF', '#FF9F40', '#FF6384', '#C9CBCF'
            ]),
            'borderWidth': 2,
            'borderColor': '#fff'
        }]
        
        return {
            'labels': labels,
            'datasets': datasets,
            'metadata': {
                'total_segments': len(labels),
                'chart_type': 'pie'
            }
        }
    
    def _process_area_chart_data(self, data: List[Dict], config: Dict) -> Dict:
        """Process data for area charts"""
        # Similar to line chart but with fill
        result = self._process_line_chart_data(data, config)
        
        # Add fill property to datasets
        for dataset in result['datasets']:
            dataset['fill'] = True
            dataset['backgroundColor'] = dataset['borderColor'] + '40'
        
        result['metadata']['chart_type'] = 'area'
        return result
    
    def _process_generic_chart_data(self, data: List[Dict], config: Dict) -> Dict:
        """Generic data processing"""
        return {
            'labels': [],
            'datasets': [],
            'raw_data': data,
            'metadata': {
                'chart_type': 'generic',
                'total_records': len(data)
            }
        }

class ChartExportService:
    """Service for exporting charts in various formats"""
    
    def __init__(self):
        self.export_dir = getattr(settings, 'CHART_EXPORT_DIR', 'exports/charts/')
    
    def export_chart(self, chart: Chart, format: str, options: Dict = None) -> str:
        """Export chart in specified format"""
        options = options or {}
        
        if format == 'png':
            return self._export_png(chart, options)
        elif format == 'pdf':
            return self._export_pdf(chart, options)
        elif format == 'svg':
            return self._export_svg(chart, options)
        elif format == 'excel':
            return self._export_excel(chart, options)
        elif format == 'csv':
            return self._export_csv(chart, options)
        elif format == 'json':
            return self._export_json(chart, options)
        else:
            raise ValueError(f"Unsupported export format: {format}")
    
    def _export_png(self, chart: Chart, options: Dict) -> str:
        """Export chart as PNG image"""
        # This would typically use a headless browser or chart rendering library
        # For now, we'll create a placeholder implementation
        
        width = options.get('width', 800)
        height = options.get('height', 600)
        
        # Create a simple chart image using matplotlib
        plt.figure(figsize=(width/100, height/100))
        
        # Get chart data
        data_service = ChartDataService()
        chart_data = data_service.get_chart_data(chart)
        
        if chart.chart_type == 'line':
            for dataset in chart_data['datasets']:
                plt.plot(chart_data['labels'], dataset['data'], label=dataset['label'])
            plt.legend()
        elif chart.chart_type == 'bar':
            dataset = chart_data['datasets'][0]
            plt.bar(chart_data['labels'], dataset['data'])
        elif chart.chart_type == 'pie':
            dataset = chart_data['datasets'][0]
            plt.pie(dataset['data'], labels=chart_data['labels'], autopct='%1.1f%%')
        
        plt.title(chart.title)
        
        # Save to file
        filename = f"{chart.id}_{int(datetime.now().timestamp())}.png"
        filepath = f"{self.export_dir}{filename}"
        plt.savefig(filepath, dpi=100, bbox_inches='tight')
        plt.close()
        
        return filepath
    
    def _export_pdf(self, chart: Chart, options: Dict) -> str:
        """Export chart as PDF"""
        filename = f"{chart.id}_{int(datetime.now().timestamp())}.pdf"
        filepath = f"{self.export_dir}{filename}"
        
        # Create PDF with chart
        c = canvas.Canvas(filepath, pagesize=letter)
        width, height = letter
        
        # Add title
        c.setFont("Helvetica-Bold", 16)
        c.drawString(50, height - 50, chart.title)
        
        # Add chart description
        if chart.description:
            c.setFont("Helvetica", 12)
            c.drawString(50, height - 80, chart.description)
        
        # Export chart as image first, then embed in PDF
        png_path = self._export_png(chart, options)
        c.drawImage(png_path, 50, height - 500, width=500, height=300)
        
        c.save()
        return filepath
    
    def _export_svg(self, chart: Chart, options: Dict) -> str:
        """Export chart as SVG"""
        filename = f"{chart.id}_{int(datetime.now().timestamp())}.svg"
        filepath = f"{self.export_dir}{filename}"
        
        # Create SVG content
        svg_content = f"""<?xml version="1.0" encoding="UTF-8"?>
        <svg width="800" height="600" xmlns="http://www.w3.org/2000/svg">
            <title>{chart.title}</title>
            <text x="50" y="30" font-family="Arial" font-size="16" font-weight="bold">{chart.title}</text>
            <!-- Chart content would be generated here -->
        </svg>"""
        
        with open(filepath, 'w') as f:
            f.write(svg_content)
        
        return filepath
    
    def _export_excel(self, chart: Chart, options: Dict) -> str:
        """Export chart data as Excel file"""
        filename = f"{chart.id}_{int(datetime.now().timestamp())}.xlsx"
        filepath = f"{self.export_dir}{filename}"
        
        # Get chart data
        data_service = ChartDataService()
        chart_data = data_service.get_chart_data(chart)
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Chart Data"
        
        # Add headers
        headers = ['Label'] + [dataset['label'] for dataset in chart_data['datasets']]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Add data
        for row, label in enumerate(chart_data['labels'], 2):
            ws.cell(row=row, column=1, value=label)
            for col, dataset in enumerate(chart_data['datasets'], 2):
                value = dataset['data'][row-2] if row-2 < len(dataset['data']) else 0
                ws.cell(row=row, column=col, value=value)
        
        wb.save(filepath)
        return filepath
    
    def _export_csv(self, chart: Chart, options: Dict) -> str:
        """Export chart data as CSV"""
        filename = f"{chart.id}_{int(datetime.now().timestamp())}.csv"
        filepath = f"{self.export_dir}{filename}"
        
        # Get chart data
        data_service = ChartDataService()
        chart_data = data_service.get_chart_data(chart)
        
        # Create DataFrame
        data_dict = {'Label': chart_data['labels']}
        for dataset in chart_data['datasets']:
            data_dict[dataset['label']] = dataset['data']
        
        df = pd.DataFrame(data_dict)
        df.to_csv(filepath, index=False)
        
        return filepath
    
    def _export_json(self, chart: Chart, options: Dict) -> str:
        """Export chart data as JSON"""
        filename = f"{chart.id}_{int(datetime.now().timestamp())}.json"
        filepath = f"{self.export_dir}{filename}"
        
        # Get chart data
        data_service = ChartDataService()
        chart_data = data_service.get_chart_data(chart)
        
        # Add chart metadata
        export_data = {
            'chart': {
                'id': str(chart.id),
                'title': chart.title,
                'type': chart.chart_type,
                'created_at': chart.created_at.isoformat(),
            },
            'data': chart_data,
            'exported_at': datetime.now().isoformat()
        }
        
        with open(filepath, 'w') as f:
            json.dump(export_data, f, indent=2)
        
        return filepath

class ChartPerformanceService:
    """Service for monitoring chart performance"""
    
    def record_performance(self, chart: Chart, load_time: float, 
                         data_size: int, render_time: float, 
                         user_agent: str, ip_address: str):
        """Record chart performance metrics"""
        ChartPerformanceMetric.objects.create(
            chart=chart,
            load_time=load_time,
            data_size=data_size,
            render_time=render_time,
            user_agent=user_agent,
            ip_address=ip_address
        )
    
    def get_performance_stats(self, chart: Chart, days: int = 30) -> Dict:
        """Get performance statistics for a chart"""
        since = timezone.now() - timedelta(days=days)
        
        metrics = ChartPerformanceMetric.objects.filter(
            chart=chart,
            timestamp__gte=since
        )
        
        if not metrics.exists():
            return {}
        
        stats = metrics.aggregate(
            avg_load_time=Avg('load_time'),
            max_load_time=Max('load_time'),
            min_load_time=Min('load_time'),
            avg_data_size=Avg('data_size'),
            avg_render_time=Avg('render_time'),
            total_views=Count('id')
        )
        
        return stats

class ChartAnalyticsService:
    """Service for chart analytics and insights"""
    
    def get_chart_insights(self, chart: Chart) -> Dict:
        """Get insights and recommendations for a chart"""
        insights = {
            'performance': self._analyze_performance(chart),
            'usage': self._analyze_usage(chart),
            'data_quality': self._analyze_data_quality(chart),
            'recommendations': []
        }
        
        # Generate recommendations based on analysis
        if insights['performance']['avg_load_time'] > 5000:  # 5 seconds
            insights['recommendations'].append({
                'type': 'performance',
                'message': 'Chart load time is high. Consider data caching or pagination.',
                'priority': 'high'
            })
        
        if insights['usage']['access_count'] < 10:
            insights['recommendations'].append({
                'type': 'usage',
                'message': 'Chart has low usage. Consider promoting or improving visibility.',
                'priority': 'medium'
            })
        
        return insights
    
    def _analyze_performance(self, chart: Chart) -> Dict:
        """Analyze chart performance"""
        perf_service = ChartPerformanceService()
        return perf_service.get_performance_stats(chart)
    
    def _analyze_usage(self, chart: Chart) -> Dict:
        """Analyze chart usage patterns"""
        return {
            'access_count': chart.access_count,
            'last_accessed': chart.last_accessed,
            'shares_count': chart.shares.count(),
            'comments_count': chart.comments.count()
        }
    
    def _analyze_data_quality(self, chart: Chart) -> Dict:
        """Analyze chart data quality"""
        data_service = ChartDataService()
        chart_data = data_service.get_chart_data(chart)
        
        total_points = len(chart_data.get('labels', []))
        null_points = 0
        
        for dataset in chart_data.get('datasets', []):
            for value in dataset.get('data', []):
                if value is None or value == 0:
                    null_points += 1
        
        return {
            'total_data_points': total_points,
            'null_data_points': null_points,
            'data_completeness': (total_points - null_points) / total_points if total_points > 0 else 0
        }
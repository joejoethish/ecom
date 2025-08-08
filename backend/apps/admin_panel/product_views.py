"""
Views for the comprehensive product management system.
"""
import csv
import json
import xml.etree.ElementTree as ET
try:
    import openpyxl
except ImportError:
    openpyxl = None
from io import StringIO, BytesIO
from decimal import Decimal
from django.db import transaction
from django.db.models import Q, Count, Avg, Sum, F, Case, When
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.core.paginator import Paginator
from django.core.files.base import ContentFile
from django.core.exceptions import ValidationError
from django.utils.text import slugify
from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django_filters.rest_framework import DjangoFilterBackend
from django_filters import rest_framework as filters

from apps.products.models import Product, Category, ProductImage
from apps.inventory.models import Inventory
from .models import AdminUser

from .product_models import (
    ProductVariant, ProductAttribute, ProductAttributeValue, ProductBundle,
    ProductBundleItem, ProductRelationship, ProductLifecycle, ProductVersion,
    ProductTemplate, ProductQuality, ProductWarranty, ProductDigitalAsset,
    ProductSyndication, ProductAnalytics
)
from .product_serializers import (
    ComprehensiveProductSerializer, ProductCreateUpdateSerializer,
    ProductVariantSerializer, ProductBundleSerializer, ProductRelationshipSerializer,
    ProductLifecycleSerializer, ProductVersionSerializer, ProductTemplateSerializer,
    ProductQualitySerializer, ProductWarrantySerializer, ProductDigitalAssetSerializer,
    ProductSyndicationSerializer, ProductAnalyticsSerializer, BulkProductOperationSerializer,
    ProductImportSerializer, ProductExportSerializer, ProductComparisonSerializer,
    ProductDuplicateDetectionSerializer, ProductAttributeSerializer
)


class ProductFilter(filters.FilterSet):
    """Advanced filtering for products."""
    
    name = filters.CharFilter(lookup_expr='icontains')
    category = filters.ModelChoiceFilter(queryset=Category.objects.all())
    brand = filters.CharFilter(lookup_expr='icontains')
    price_min = filters.NumberFilter(field_name='price', lookup_expr='gte')
    price_max = filters.NumberFilter(field_name='price', lookup_expr='lte')
    is_active = filters.BooleanFilter()
    is_featured = filters.BooleanFilter()
    status = filters.ChoiceFilter(choices=[
        ('draft', 'Draft'),
        ('active', 'Active'),
        ('inactive', 'Inactive'),
        ('out_of_stock', 'Out of Stock'),
    ])
    created_after = filters.DateTimeFilter(field_name='created_at', lookup_expr='gte')
    created_before = filters.DateTimeFilter(field_name='created_at', lookup_expr='lte')
    has_variants = filters.BooleanFilter(method='filter_has_variants')
    low_stock = filters.BooleanFilter(method='filter_low_stock')
    lifecycle_stage = filters.CharFilter(field_name='lifecycle__current_stage')
    
    class Meta:
        model = Product
        fields = [
            'name', 'category', 'brand', 'price_min', 'price_max',
            'is_active', 'is_featured', 'status', 'created_after',
            'created_before', 'has_variants', 'low_stock', 'lifecycle_stage'
        ]
    
    def filter_has_variants(self, queryset, name, value):
        """Filter products that have variants."""
        if value:
            return queryset.filter(variants__isnull=False).distinct()
        return queryset.filter(variants__isnull=True)
    
    def filter_low_stock(self, queryset, name, value):
        """Filter products with low stock."""
        if value:
            return queryset.filter(
                inventory__quantity__lte=F('inventory__reorder_point')
            )
        return queryset


class ComprehensiveProductViewSet(viewsets.ModelViewSet):
    """
    Comprehensive product management with advanced features.
    
    Provides CRUD operations, bulk operations, import/export,
    variant management, and analytics.
    """
    
    queryset = Product.objects.select_related(
        'category', 'brand', 'lifecycle', 'quality', 'warranty', 'analytics'
    ).prefetch_related(
        'images', 'variants', 'relationships_from', 'digital_assets',
        'syndications', 'inventory'
    ).all()
    
    serializer_class = ComprehensiveProductSerializer
    permission_classes = [permissions.IsAuthenticated]  # Will be updated with proper admin permissions
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    filter_backends = [DjangoFilterBackend]
    filterset_class = ProductFilter
    
    def get_serializer_class(self):
        """Return appropriate serializer based on action."""
        if self.action in ['create', 'update', 'partial_update']:
            return ProductCreateUpdateSerializer
        return ComprehensiveProductSerializer
    
    def get_queryset(self):
        """Get filtered and optimized queryset."""
        queryset = super().get_queryset()
        
        # Add computed fields
        queryset = queryset.annotate(
            effective_price=Case(
                When(discount_price__isnull=False, then=F('discount_price')),
                default=F('price')
            ),
            discount_percentage=Case(
                When(discount_price__isnull=False, 
                     then=(F('price') - F('discount_price')) * 100 / F('price')),
                default=0
            ),
            average_rating=Avg('reviews__rating'),
            total_reviews=Count('reviews'),
            total_sales=Sum('order_items__quantity'),
            total_revenue=Sum(F('order_items__quantity') * F('order_items__price'))
        )
        
        return queryset
    
    @action(detail=False, methods=['post'])
    def bulk_operations(self, request):
        """
        Perform bulk operations on multiple products.
        
        Supported operations:
        - activate/deactivate
        - feature/unfeature
        - delete
        - update_category
        - update_status
        - export
        """
        serializer = BulkProductOperationSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        product_ids = serializer.validated_data['product_ids']
        operation = serializer.validated_data['operation']
        parameters = serializer.validated_data.get('parameters', {})
        
        products = Product.objects.filter(id__in=product_ids)
        
        if not products.exists():
            return Response(
                {'error': 'No products found with provided IDs'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        try:
            with transaction.atomic():
                if operation == 'activate':
                    products.update(is_active=True)
                    message = f"Activated {products.count()} products"
                
                elif operation == 'deactivate':
                    products.update(is_active=False)
                    message = f"Deactivated {products.count()} products"
                
                elif operation == 'feature':
                    products.update(is_featured=True)
                    message = f"Featured {products.count()} products"
                
                elif operation == 'unfeature':
                    products.update(is_featured=False)
                    message = f"Unfeatured {products.count()} products"
                
                elif operation == 'delete':
                    count = products.count()
                    products.delete()
                    message = f"Deleted {count} products"
                
                elif operation == 'update_category':
                    category_id = parameters.get('category_id')
                    if not category_id:
                        return Response(
                            {'error': 'category_id required for update_category operation'},
                            status=status.HTTP_400_BAD_REQUEST
                        )
                    products.update(category_id=category_id)
                    message = f"Updated category for {products.count()} products"
                
                elif operation == 'update_status':
                    new_status = parameters.get('status')
                    if not new_status:
                        return Response(
                            {'error': 'status required for update_status operation'},
                            status=status.HTTP_400_BAD_REQUEST
                        )
                    products.update(status=new_status)
                    message = f"Updated status for {products.count()} products"
                
                elif operation == 'export':
                    return self._export_products(products, parameters)
                
                else:
                    return Response(
                        {'error': f'Unknown operation: {operation}'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            
            return Response({
                'success': True,
                'message': message,
                'affected_count': len(product_ids)
            })
        
        except Exception as e:
            return Response(
                {'error': f'Bulk operation failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['post'])
    def import_products(self, request):
        """
        Import products from CSV, Excel, JSON, or XML files.
        
        Supports:
        - Multiple file formats
        - Bulk creation and updates
        - Category creation
        - Validation and error reporting
        """
        serializer = ProductImportSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        file = serializer.validated_data['file']
        file_format = serializer.validated_data['file_format']
        import_options = serializer.validated_data.get('import_options', {})
        update_existing = serializer.validated_data.get('update_existing', False)
        create_categories = serializer.validated_data.get('create_categories', False)
        
        try:
            if file_format == 'csv':
                result = self._import_from_csv(file, import_options, update_existing, create_categories)
            elif file_format == 'excel':
                result = self._import_from_excel(file, import_options, update_existing, create_categories)
            elif file_format == 'json':
                result = self._import_from_json(file, import_options, update_existing, create_categories)
            elif file_format == 'xml':
                result = self._import_from_xml(file, import_options, update_existing, create_categories)
            else:
                return Response(
                    {'error': f'Unsupported file format: {file_format}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            return Response(result)
        
        except Exception as e:
            return Response(
                {'error': f'Import failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['post'])
    def export_products(self, request):
        """
        Export products to various formats.
        
        Supports CSV, Excel, JSON, and XML formats with
        customizable field selection and filtering.
        """
        serializer = ProductExportSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        product_ids = serializer.validated_data.get('product_ids')
        export_format = serializer.validated_data['export_format']
        include_images = serializer.validated_data.get('include_images', False)
        include_variants = serializer.validated_data.get('include_variants', False)
        include_relationships = serializer.validated_data.get('include_relationships', False)
        filters = serializer.validated_data.get('filters', {})
        
        # Get products to export
        if product_ids:
            products = Product.objects.filter(id__in=product_ids)
        else:
            products = self.filter_queryset(self.get_queryset())
            # Apply additional filters
            for field, value in filters.items():
                if hasattr(Product, field):
                    products = products.filter(**{field: value})
        
        return self._export_products(products, {
            'format': export_format,
            'include_images': include_images,
            'include_variants': include_variants,
            'include_relationships': include_relationships
        })
    
    @action(detail=False, methods=['post'])
    def compare_products(self, request):
        """
        Compare multiple products side by side.
        
        Returns detailed comparison data for specified products
        and fields.
        """
        serializer = ProductComparisonSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        product_ids = serializer.validated_data['product_ids']
        comparison_fields = serializer.validated_data['comparison_fields']
        
        products = Product.objects.filter(id__in=product_ids).select_related(
            'category', 'brand'
        ).prefetch_related('images', 'variants')
        
        if not products.exists():
            return Response(
                {'error': 'No products found with provided IDs'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        comparison_data = []
        for product in products:
            product_data = {'id': product.id}
            
            for field in comparison_fields:
                if field == 'name':
                    product_data['name'] = product.name
                elif field == 'price':
                    product_data['price'] = float(product.price)
                    product_data['discount_price'] = float(product.discount_price) if product.discount_price else None
                elif field == 'category':
                    product_data['category'] = product.category.name if product.category else None
                elif field == 'brand':
                    product_data['brand'] = product.brand.name if product.brand else None
                elif field == 'rating':
                    product_data['rating'] = product.average_rating
                elif field == 'stock':
                    try:
                        product_data['stock'] = product.inventory.quantity
                    except:
                        product_data['stock'] = 0
                elif field == 'variants':
                    product_data['variants_count'] = product.variants.count()
                elif field == 'images':
                    product_data['images_count'] = product.images.count()
                elif hasattr(product, field):
                    product_data[field] = getattr(product, field)
            
            comparison_data.append(product_data)
        
        return Response({
            'products': comparison_data,
            'comparison_fields': comparison_fields,
            'total_products': len(comparison_data)
        })
    
    @action(detail=False, methods=['post'])
    def detect_duplicates(self, request):
        """
        Detect duplicate products based on various criteria.
        
        Uses name similarity, SKU matching, and other criteria
        to identify potential duplicates.
        """
        serializer = ProductDuplicateDetectionSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        detection_criteria = serializer.validated_data['detection_criteria']
        auto_merge = serializer.validated_data.get('auto_merge', False)
        merge_strategy = serializer.validated_data.get('merge_strategy', 'manual')
        
        # This is a simplified implementation
        # In a real system, you'd use more sophisticated algorithms
        duplicates = []
        products = Product.objects.all()
        
        # Group by potential duplicate criteria
        if detection_criteria.get('sku_match', True):
            sku_groups = {}
            for product in products:
                if product.sku:
                    if product.sku not in sku_groups:
                        sku_groups[product.sku] = []
                    sku_groups[product.sku].append(product)
            
            for sku, group in sku_groups.items():
                if len(group) > 1:
                    duplicates.append({
                        'type': 'sku_match',
                        'criteria': sku,
                        'products': [{'id': p.id, 'name': p.name} for p in group],
                        'confidence': 1.0
                    })
        
        # Name similarity detection would require additional libraries
        # like difflib or fuzzywuzzy for proper implementation
        
        return Response({
            'duplicates': duplicates,
            'total_duplicate_groups': len(duplicates),
            'auto_merge_enabled': auto_merge,
            'merge_strategy': merge_strategy
        })
    
    @action(detail=True, methods=['get'])
    def analytics(self, request, pk=None):
        """
        Get detailed analytics for a specific product.
        
        Returns performance metrics, sales data, and insights.
        """
        product = self.get_object()
        
        try:
            analytics = product.analytics
            analytics.calculate_metrics()  # Update metrics
            
            serializer = ProductAnalyticsSerializer(analytics)
            return Response(serializer.data)
        
        except ProductAnalytics.DoesNotExist:
            # Create analytics if doesn't exist
            analytics = ProductAnalytics.objects.create(product=product)
            serializer = ProductAnalyticsSerializer(analytics)
            return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def lifecycle_transition(self, request, pk=None):
        """
        Transition product to a new lifecycle stage.
        
        Handles workflow validation and approval requirements.
        """
        product = self.get_object()
        new_stage = request.data.get('stage')
        notes = request.data.get('notes', '')
        
        if not new_stage:
            return Response(
                {'error': 'stage is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            lifecycle, created = ProductLifecycle.objects.get_or_create(
                product=product,
                defaults={'current_stage': 'draft'}
            )
            
            lifecycle.transition_to_stage(new_stage, request.user, notes)
            
            serializer = ProductLifecycleSerializer(lifecycle)
            return Response({
                'success': True,
                'message': f'Product transitioned to {new_stage}',
                'lifecycle': serializer.data
            })
        
        except ValueError as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['post'])
    def create_version(self, request, pk=None):
        """
        Create a new version of the product.
        
        Captures current state and creates version history.
        """
        product = self.get_object()
        version_number = request.data.get('version_number')
        change_summary = request.data.get('change_summary', '')
        
        if not version_number:
            return Response(
                {'error': 'version_number is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check if version already exists
        if ProductVersion.objects.filter(product=product, version_number=version_number).exists():
            return Response(
                {'error': f'Version {version_number} already exists'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Create version snapshot
        product_data = ComprehensiveProductSerializer(product).data
        variant_data = list(product.variants.values())
        image_data = list(product.images.values())
        
        version = ProductVersion.objects.create(
            product=product,
            version_number=version_number,
            product_data=product_data,
            variant_data=variant_data,
            image_data=image_data,
            change_summary=change_summary,
            changed_by=request.user,
            is_current=True
        )
        
        serializer = ProductVersionSerializer(version)
        return Response({
            'success': True,
            'message': f'Version {version_number} created',
            'version': serializer.data
        })
    
    def _import_from_csv(self, file, options, update_existing, create_categories):
        """Import products from CSV file."""
        content = file.read().decode('utf-8')
        csv_reader = csv.DictReader(StringIO(content))
        
        created_count = 0
        updated_count = 0
        errors = []
        
        with transaction.atomic():
            for row_num, row in enumerate(csv_reader, start=2):
                try:
                    # Process each row
                    product_data = self._process_csv_row(row, create_categories)
                    
                    if update_existing and product_data.get('sku'):
                        product, created = Product.objects.update_or_create(
                            sku=product_data['sku'],
                            defaults=product_data
                        )
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                    else:
                        Product.objects.create(**product_data)
                        created_count += 1
                
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
        
        return {
            'success': True,
            'created_count': created_count,
            'updated_count': updated_count,
            'errors': errors
        }
    
    def _import_from_excel(self, file, options, update_existing, create_categories):
        """Import products from Excel file."""
        workbook = openpyxl.load_workbook(file)
        worksheet = workbook.active
        
        # Get headers from first row
        headers = [cell.value for cell in worksheet[1]]
        
        created_count = 0
        updated_count = 0
        errors = []
        
        with transaction.atomic():
            for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Create dict from headers and row values
                    row_data = dict(zip(headers, row))
                    product_data = self._process_excel_row(row_data, create_categories)
                    
                    if update_existing and product_data.get('sku'):
                        product, created = Product.objects.update_or_create(
                            sku=product_data['sku'],
                            defaults=product_data
                        )
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                    else:
                        Product.objects.create(**product_data)
                        created_count += 1
                
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
        
        return {
            'success': True,
            'created_count': created_count,
            'updated_count': updated_count,
            'errors': errors
        }
    
    def _import_from_json(self, file, options, update_existing, create_categories):
        """Import products from JSON file."""
        content = file.read().decode('utf-8')
        data = json.loads(content)
        
        if not isinstance(data, list):
            data = [data]
        
        created_count = 0
        updated_count = 0
        errors = []
        
        with transaction.atomic():
            for index, item in enumerate(data):
                try:
                    product_data = self._process_json_item(item, create_categories)
                    
                    if update_existing and product_data.get('sku'):
                        product, created = Product.objects.update_or_create(
                            sku=product_data['sku'],
                            defaults=product_data
                        )
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                    else:
                        Product.objects.create(**product_data)
                        created_count += 1
                
                except Exception as e:
                    errors.append(f"Item {index + 1}: {str(e)}")
        
        return {
            'success': True,
            'created_count': created_count,
            'updated_count': updated_count,
            'errors': errors
        }
    
    def _import_from_xml(self, file, options, update_existing, create_categories):
        """Import products from XML file."""
        content = file.read().decode('utf-8')
        root = ET.fromstring(content)
        
        created_count = 0
        updated_count = 0
        errors = []
        
        with transaction.atomic():
            for index, product_elem in enumerate(root.findall('.//product')):
                try:
                    product_data = self._process_xml_element(product_elem, create_categories)
                    
                    if update_existing and product_data.get('sku'):
                        product, created = Product.objects.update_or_create(
                            sku=product_data['sku'],
                            defaults=product_data
                        )
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                    else:
                        Product.objects.create(**product_data)
                        created_count += 1
                
                except Exception as e:
                    errors.append(f"Product {index + 1}: {str(e)}")
        
        return {
            'success': True,
            'created_count': created_count,
            'updated_count': updated_count,
            'errors': errors
        }
    
    def _process_csv_row(self, row, create_categories):
        """Process a CSV row into product data."""
        product_data = {
            'name': row.get('name', ''),
            'description': row.get('description', ''),
            'sku': row.get('sku', ''),
            'price': Decimal(row.get('price', '0')),
            'is_active': row.get('is_active', 'true').lower() == 'true'
        }
        
        # Handle category
        category_name = row.get('category')
        if category_name:
            if create_categories:
                category, _ = Category.objects.get_or_create(
                    name=category_name,
                    defaults={'slug': slugify(category_name)}
                )
                product_data['category'] = category
            else:
                try:
                    product_data['category'] = Category.objects.get(name=category_name)
                except Category.DoesNotExist:
                    pass
        
        return product_data
    
    def _process_excel_row(self, row_data, create_categories):
        """Process an Excel row into product data."""
        return self._process_csv_row(row_data, create_categories)
    
    def _process_json_item(self, item, create_categories):
        """Process a JSON item into product data."""
        product_data = {
            'name': item.get('name', ''),
            'description': item.get('description', ''),
            'sku': item.get('sku', ''),
            'price': Decimal(str(item.get('price', '0'))),
            'is_active': item.get('is_active', True)
        }
        
        # Handle category
        category_name = item.get('category')
        if category_name:
            if create_categories:
                category, _ = Category.objects.get_or_create(
                    name=category_name,
                    defaults={'slug': slugify(category_name)}
                )
                product_data['category'] = category
            else:
                try:
                    product_data['category'] = Category.objects.get(name=category_name)
                except Category.DoesNotExist:
                    pass
        
        return product_data
    
    def _process_xml_element(self, element, create_categories):
        """Process an XML element into product data."""
        product_data = {
            'name': element.findtext('name', ''),
            'description': element.findtext('description', ''),
            'sku': element.findtext('sku', ''),
            'price': Decimal(element.findtext('price', '0')),
            'is_active': element.findtext('is_active', 'true').lower() == 'true'
        }
        
        # Handle category
        category_name = element.findtext('category')
        if category_name:
            if create_categories:
                category, _ = Category.objects.get_or_create(
                    name=category_name,
                    defaults={'slug': slugify(category_name)}
                )
                product_data['category'] = category
            else:
                try:
                    product_data['category'] = Category.objects.get(name=category_name)
                except Category.DoesNotExist:
                    pass
        
        return product_data
    
    def _export_products(self, products, options):
        """Export products to specified format."""
        export_format = options.get('format', 'csv')
        include_images = options.get('include_images', False)
        include_variants = options.get('include_variants', False)
        include_relationships = options.get('include_relationships', False)
        
        if export_format == 'csv':
            return self._export_to_csv(products, options)
        elif export_format == 'excel':
            return self._export_to_excel(products, options)
        elif export_format == 'json':
            return self._export_to_json(products, options)
        elif export_format == 'xml':
            return self._export_to_xml(products, options)
        else:
            return Response(
                {'error': f'Unsupported export format: {export_format}'},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    def _export_to_csv(self, products, options):
        """Export products to CSV format."""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="products.csv"'
        
        writer = csv.writer(response)
        
        # Write headers
        headers = [
            'id', 'name', 'sku', 'description', 'category', 'brand',
            'price', 'discount_price', 'is_active', 'is_featured',
            'created_at', 'updated_at'
        ]
        writer.writerow(headers)
        
        # Write data
        for product in products:
            writer.writerow([
                product.id,
                product.name,
                product.sku,
                product.description,
                product.category.name if product.category else '',
                product.brand.name if product.brand else '',
                product.price,
                product.discount_price or '',
                product.is_active,
                product.is_featured,
                product.created_at,
                product.updated_at
            ])
        
        return response
    
    def _export_to_excel(self, products, options):
        """Export products to Excel format."""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Products"
        
        # Write headers
        headers = [
            'ID', 'Name', 'SKU', 'Description', 'Category', 'Brand',
            'Price', 'Discount Price', 'Active', 'Featured',
            'Created At', 'Updated At'
        ]
        worksheet.append(headers)
        
        # Write data
        for product in products:
            worksheet.append([
                product.id,
                product.name,
                product.sku,
                product.description,
                product.category.name if product.category else '',
                product.brand.name if product.brand else '',
                float(product.price),
                float(product.discount_price) if product.discount_price else None,
                product.is_active,
                product.is_featured,
                product.created_at,
                product.updated_at
            ])
        
        # Save to BytesIO
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="products.xlsx"'
        
        return response
    
    def _export_to_json(self, products, options):
        """Export products to JSON format."""
        serializer = ComprehensiveProductSerializer(products, many=True)
        
        response = HttpResponse(
            json.dumps(serializer.data, indent=2, default=str),
            content_type='application/json'
        )
        response['Content-Disposition'] = 'attachment; filename="products.json"'
        
        return response
    
    def _export_to_xml(self, products, options):
        """Export products to XML format."""
        root = ET.Element("products")
        
        for product in products:
            product_elem = ET.SubElement(root, "product")
            product_elem.set("id", str(product.id))
            
            ET.SubElement(product_elem, "name").text = product.name
            ET.SubElement(product_elem, "sku").text = product.sku
            ET.SubElement(product_elem, "description").text = product.description or ''
            ET.SubElement(product_elem, "category").text = product.category.name if product.category else ''
            ET.SubElement(product_elem, "brand").text = product.brand.name if product.brand else ''
            ET.SubElement(product_elem, "price").text = str(product.price)
            ET.SubElement(product_elem, "discount_price").text = str(product.discount_price) if product.discount_price else ''
            ET.SubElement(product_elem, "is_active").text = str(product.is_active).lower()
            ET.SubElement(product_elem, "is_featured").text = str(product.is_featured).lower()
            ET.SubElement(product_elem, "created_at").text = product.created_at.isoformat()
            ET.SubElement(product_elem, "updated_at").text = product.updated_at.isoformat()
        
        xml_string = ET.tostring(root, encoding='unicode')
        
        response = HttpResponse(xml_string, content_type='application/xml')
        response['Content-Disposition'] = 'attachment; filename="products.xml"'
        
        return response


class ProductVariantViewSet(viewsets.ModelViewSet):
    """
    Product variant management with attributes and SKU generation.
    """
    
    queryset = ProductVariant.objects.select_related('product').all()
    serializer_class = ProductVariantSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['product', 'is_active', 'is_default']
    
    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        """Create multiple variants for a product."""
        product_id = request.data.get('product_id')
        variants_data = request.data.get('variants', [])
        
        if not product_id or not variants_data:
            return Response(
                {'error': 'product_id and variants data required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            product = Product.objects.get(id=product_id)
        except Product.DoesNotExist:
            return Response(
                {'error': 'Product not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        created_variants = []
        errors = []
        
        with transaction.atomic():
            for variant_data in variants_data:
                try:
                    variant_data['product'] = product.id
                    serializer = ProductVariantSerializer(data=variant_data)
                    serializer.is_valid(raise_exception=True)
                    variant = serializer.save()
                    created_variants.append(serializer.data)
                except Exception as e:
                    errors.append(f"Variant {variant_data.get('name', 'Unknown')}: {str(e)}")
        
        return Response({
            'success': True,
            'created_count': len(created_variants),
            'variants': created_variants,
            'errors': errors
        })


class ProductAttributeViewSet(viewsets.ModelViewSet):
    """
    Product attribute management for variant configuration.
    """
    
    queryset = ProductAttribute.objects.prefetch_related('values').all()
    serializer_class = ProductAttributeSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['attribute_type', 'is_required', 'is_variant_attribute']


class ProductBundleViewSet(viewsets.ModelViewSet):
    """
    Product bundle and kit management with dynamic pricing.
    """
    
    queryset = ProductBundle.objects.prefetch_related('items__product', 'items__variant').all()
    serializer_class = ProductBundleSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['bundle_type', 'pricing_type', 'is_active']


class ProductTemplateViewSet(viewsets.ModelViewSet):
    """
    Product template system for quick product creation.
    """
    
    queryset = ProductTemplate.objects.select_related('category', 'created_by').all()
    serializer_class = ProductTemplateSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['category', 'is_active', 'created_by']
    
    @action(detail=True, methods=['post'])
    def create_product(self, request, pk=None):
        """Create a new product using this template."""
        template = self.get_object()
        product_data = request.data.get('product_data', {})
        
        try:
            product = template.create_product_from_template(product_data, request.user)
            serializer = ComprehensiveProductSerializer(product)
            
            return Response({
                'success': True,
                'message': f'Product created from template: {template.name}',
                'product': serializer.data
            })
        
        except Exception as e:
            return Response(
                {'error': f'Failed to create product from template: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ProductAnalyticsViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Product analytics and performance metrics.
    """
    
    queryset = ProductAnalytics.objects.select_related('product').all()
    serializer_class = ProductAnalyticsSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    @action(detail=False, methods=['get'])
    def dashboard(self, request):
        """Get analytics dashboard data."""
        # Top performing products
        top_products = ProductAnalytics.objects.select_related('product').order_by(
            '-total_revenue'
        )[:10]
        
        # Category performance
        category_performance = ProductAnalytics.objects.values(
            'product__category__name'
        ).annotate(
            total_revenue=Sum('total_revenue'),
            total_sales=Sum('total_sales'),
            avg_conversion=Avg('conversion_rate')
        ).order_by('-total_revenue')[:10]
        
        # Overall metrics
        overall_metrics = ProductAnalytics.objects.aggregate(
            total_products=Count('id'),
            total_revenue=Sum('total_revenue'),
            total_sales=Sum('total_sales'),
            avg_conversion=Avg('conversion_rate')
        )
        
        return Response({
            'top_products': ProductAnalyticsSerializer(top_products, many=True).data,
            'category_performance': list(category_performance),
            'overall_metrics': overall_metrics
        })